"""Import CSV/Excel de contacts en masse (staff only)."""
from __future__ import annotations

import io
import secrets
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel

from albarka_auth import require_staff
from albarka_contacts import CONTACT_CATEGORIES, CONTACT_FUNCTIONS, CONTACT_SCOPES
from db import db

router = APIRouter(prefix="/contacts", tags=["Contacts"])


CSV_HEADERS = [
    "full_name", "function", "organization", "email", "phone",
    "is_primary", "can_receive_notifications", "channels", "categories", "notes",
]


def _parse_bool(v: str) -> bool:
    return str(v).strip().lower() in ("1", "true", "vrai", "oui", "yes", "y", "x")


def _split(v: str) -> List[str]:
    return [x.strip() for x in str(v).split("|") if x.strip()]


def _parse_csv(data: bytes) -> List[dict]:
    import csv
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _parse_xlsx(data: bytes) -> List[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
    return out


class ImportReport(BaseModel):
    imported: int
    updated: int
    skipped: int
    errors: List[dict]


@router.post("/import", response_model=ImportReport)
async def import_contacts(
    file: UploadFile = File(...),
    scope: str = Form(...),
    tenant_id: Optional[str] = Form(None),
    user: dict = Depends(require_staff()),
):
    if scope not in CONTACT_SCOPES:
        raise HTTPException(status_code=400, detail=f"scope invalide : {scope}")
    resolved_tenant = tenant_id if scope == "client" else "cabinet"
    if scope == "client":
        if not resolved_tenant:
            raise HTTPException(status_code=400, detail="tenant_id requis pour scope=client")
        exists = await db.users.find_one({"id": resolved_tenant, "roles": "client"}, {"_id": 0, "id": 1})
        if not exists:
            raise HTTPException(status_code=404, detail="Client introuvable")

    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Fichier trop volumineux (5 Mo max)")
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        rows = _parse_csv(data)
    elif filename.endswith(".xlsx"):
        rows = _parse_xlsx(data)
    else:
        raise HTTPException(status_code=400, detail="Format non supporté (attendu .csv ou .xlsx)")

    imported = updated = skipped = 0
    errors: List[dict] = []
    for idx, raw in enumerate(rows, start=2):  # +2 = header row is 1
        row = {k: ("" if v is None else str(v).strip()) for k, v in raw.items()}
        full_name = row.get("full_name") or row.get("Nom complet") or ""
        email = (row.get("email") or "").lower()
        phone = row.get("phone") or ""
        if not full_name:
            skipped += 1
            errors.append({"row": idx, "reason": "full_name manquant"})
            continue
        if not email and not phone:
            skipped += 1
            errors.append({"row": idx, "reason": "email ou téléphone requis"})
            continue

        function = row.get("function", "autre") or "autre"
        if function not in CONTACT_FUNCTIONS:
            errors.append({"row": idx, "reason": f"function inconnue '{function}' → 'autre'"})
            function = "autre"

        channels_str = row.get("channels", "email") or "email"
        channels = [c for c in _split(channels_str) if c in ("email", "whatsapp")] or ["email"]

        categories_str = row.get("categories", "") or ""
        categories = [c for c in _split(categories_str) if c in CONTACT_CATEGORIES]

        # Upsert on (scope, tenant_id, email) — email is the natural key
        query = {"scope": scope, "tenant_id": resolved_tenant}
        if email:
            query["email"] = email
        else:
            query["phone"] = phone
        existing = await db.contacts.find_one(query, {"_id": 0})
        now = datetime.now(timezone.utc).isoformat()
        contact = {
            "scope": scope,
            "tenant_id": resolved_tenant,
            "full_name": full_name,
            "function": function,
            "organization": row.get("organization") or None,
            "email": email or None,
            "phone": phone or None,
            "is_primary": _parse_bool(row.get("is_primary", "")),
            "is_active": True,
            "can_receive_notifications": _parse_bool(row.get("can_receive_notifications", "true"))
                if row.get("can_receive_notifications") else True,
            "channels": channels,
            "categories": categories,
            "notes": row.get("notes") or None,
            "updated_at": now,
        }
        if existing:
            await db.contacts.update_one({"id": existing["id"]}, {"$set": contact})
            updated += 1
        else:
            contact["id"] = secrets.token_urlsafe(12)
            contact["created_at"] = now
            contact["created_by"] = user["id"]
            await db.contacts.insert_one(contact)
            imported += 1

    return ImportReport(imported=imported, updated=updated, skipped=skipped, errors=errors[:50])


@router.get("/import/template")
async def import_template(user: dict = Depends(require_staff())):
    """Renvoie un CSV modèle avec entêtes + exemple."""
    import csv
    from fastapi.responses import Response
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(CSV_HEADERS)
    w.writerow([
        "Boukary Sawadogo", "dg", "Sawadogo SARL", "boukary@sawadogo.bf",
        "+22670000010", "true", "true", "email|whatsapp", "principal", "Contact primaire DG",
    ])
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="contacts_template.csv"'},
    )
